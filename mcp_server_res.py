from mcp.server.fastmcp import FastMCP 
from pydantic import BaseModel
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from typing import List, Optional

load_dotenv()

mcp = FastMCP("Restaurant")


class MenuItem(BaseModel):
    item_id: str
    name: str
    category: str
    description: str
    price: float
    ingredients: List[str]
    dietary_tags: List[str]
    available: bool
    preparation_time: int
    popularity_score: int


class OrderItem(BaseModel):
    item_id: str
    name: str
    quantity: int
    special_requests: Optional[str] = None
    price: float


class CustomerInfo(BaseModel):
    name: str
    table_number: int
    phone: str
    email: Optional[str] = None


class Order(BaseModel):
    order_id: str
    customer: CustomerInfo
    items: List[OrderItem]
    total_amount: float
    tax_amount: float
    grand_total: float
    special_instructions: Optional[str] = None
    timestamp: str
    status: str


menu_database = {
    "app_001": MenuItem(
        item_id="app_001",
        name="Bruschetta",
        category="appetizer",
        description="Toasted bread with fresh tomatoes, basil, and garlic",
        price=8.99,
        ingredients=["bread", "tomatoes", "basil", "garlic", "olive oil"],
        dietary_tags=["vegetarian"],
        available=True,
        preparation_time=10,
        popularity_score=8
    ),
    "app_002": MenuItem(
        item_id="app_002",
        name="Buffalo Wings",
        category="appetizer",
        description="Crispy chicken wings with spicy buffalo sauce",
        price=12.99,
        ingredients=["chicken wings", "buffalo sauce", "butter", "celery"],
        dietary_tags=["spicy"],
        available=True,
        preparation_time=15,
        popularity_score=9
    ),
    "main_001": MenuItem(
        item_id="main_001",
        name="Grilled Salmon",
        category="main",
        description="Fresh Atlantic salmon with lemon butter sauce and vegetables",
        price=24.99,
        ingredients=["salmon", "lemon", "butter", "asparagus", "potatoes"],
        dietary_tags=["gluten-free"],
        available=True,
        preparation_time=25,
        popularity_score=9
    ),
    "main_002": MenuItem(
        item_id="main_002",
        name="Chicken Alfredo Pasta",
        category="main",
        description="Creamy fettuccine pasta with grilled chicken",
        price=18.99,
        ingredients=["fettuccine", "chicken", "cream", "parmesan", "garlic"],
        dietary_tags=[],
        available=True,
        preparation_time=20,
        popularity_score=10
    ),
    "des_001": MenuItem(
        item_id="des_001",
        name="Chocolate Lava Cake",
        category="dessert",
        description="Warm chocolate cake with molten center and vanilla ice cream",
        price=8.99,
        ingredients=["chocolate", "flour", "eggs", "butter", "ice cream"],
        dietary_tags=["vegetarian"],
        available=True,
        preparation_time=12,
        popularity_score=10
    ),
    "drink_001": MenuItem(
        item_id="drink_001",
        name="Fresh Lemonade",
        category="drink",
        description="Freshly squeezed lemonade with mint",
        price=4.99,
        ingredients=["lemon", "sugar", "water", "mint"],
        dietary_tags=["vegan", "vegetarian"],
        available=True,
        preparation_time=5,
        popularity_score=8
    ),
}

orders = {}


@mcp.tool()
def fetch_menu(category: str = "all"):
    """
    Fetch menu items by category.
    Categories: appetizer, main, dessert, drink, all
    """
    if category.lower() == "all":
        return list(menu_database.values())

    items = [item for item in menu_database.values()
             if item.category.lower() == category.lower()]

    if not items:
        return f"No items found in category: {category}"

    return items


@mcp.tool()
def calculate_total(items: list):
    """
    Calculate total price for items using item_ids
    """
    total = 0
    for item_id in items:
        if item_id not in menu_database:
            raise ValueError(f"Item ID '{item_id}' not found in menu.")
        total += menu_database[item_id].price

    tax = total * 0.1
    grand_total = total + tax

    return {
        "subtotal": round(total, 2),
        "tax": round(tax, 2),
        "total": round(grand_total, 2)
    }


@mcp.tool()
def create_order(customer_name: str, location: str, phone: str, items: list):
    """
    Create an order and save it
    Example: create_order("John", "Table 5", "555-1234", ["main_001", "drink_001"])
    """
    order_id = f"ORD-{len(orders) + 1}"

    for item in items:
        if item not in menu_database:
            return {"error": f"Item '{item}' not found in menu"}

    calculation = calculate_total(items)

    order = {
        "order_id": order_id,
        "customer": customer_name,
        "location": location,
        "phone": phone,
        "items": items,
        "total": calculation["total"]
    }

    orders[order_id] = order

    return {
        "success": True,
        "order_id": order_id,
        "message": f"Order created for {customer_name} at location {location}",
        "order": order
    }


@mcp.tool()
def send_to_kitchen(order_id: str, chef_email: str = "saragamilmohamed@gmail.com"):
    """Send order to kitchen via email"""
    if order_id not in orders:
        return {"error": "Order not found"}

    order = orders[order_id]

    subject = f"NEW ORDER #{order_id} - Location {order['location']}"
    body = f"""
NEW ORDER RECEIVED
==========================================

Order ID: {order_id}
Customer: {order['customer']}
Location: {order['location']}
Phone: {order['phone']}

ITEMS ORDERED:
{chr(10).join(f"  • {item}" for item in order['items'])}

TOTAL: ${order['total']}

==========================================
Please prepare this order immediately.
    """

    sender_email = os.getenv("SENDER_EMAIL")
    sender_password = os.getenv("SENDER_PASSWORD")

    try:
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = chef_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login(sender_email, sender_password)
        server.send_message(msg)
        server.quit()

        return {
            "success": True,
            "message": f"Order {order_id} sent to kitchen via email.",
            "email_sent_to": chef_email,
            "order_details": body
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"Failed to send email: {str(e)}",
            "order_details": body,
            "note": "Order saved but email failed. Check email settings in .env file."
        }


@mcp.tool()
def save_in_excel(order_id: str, customer_name: str, customer_phone_number: str, customer_location: str, items: list):
    """Save order details to Excel file"""
    try:
        excel_file = "orders.xlsx"

        if os.path.exists(excel_file):
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Orders"

            headers = ["Order ID", "Date", "Customer Name", "Phone Number", "Location", "Items"]
            sheet.append(headers)

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font

        items_str = ", ".join(items)
        date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        sheet.append([order_id, date_str, customer_name, customer_phone_number, customer_location, items_str])

        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 30

        workbook.save(excel_file)

        return {
            "success": True,
            "message": f"Order {order_id} saved to Excel.",
            "file": excel_file
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"Failed to save to Excel: {str(e)}"
        }


if __name__ == "__main__":
    mcp.run()
